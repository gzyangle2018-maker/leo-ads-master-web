"""
Leo Ads Master - Excel Exporter (v2.1)
Generates formatted Excel reports with 12-dimension analysis.
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd


class ExcelExporter:
    def __init__(self):
        # iOS-style dark accent colors
        self.header_fill = PatternFill(start_color='1C1C1E', end_color='1C1C1E', fill_type='solid')
        self.header_font = Font(color='FF6B6B', bold=True, size=11, name='Microsoft YaHei')
        self.title_font = Font(color='FFFFFF', bold=True, size=18, name='Microsoft YaHei')
        self.subtitle_font = Font(color='8E8E93', bold=True, size=12, name='Microsoft YaHei')
        self.normal_font = Font(color='1C1C1E', size=10, name='Microsoft YaHei')
        self.high_acos_font = Font(color='C00000', bold=True, size=10, name='Microsoft YaHei')
        self.low_acos_font = Font(color='00B050', bold=True, size=10, name='Microsoft YaHei')
        self.add_font = Font(color='007AFF', bold=True, size=10, name='Microsoft YaHei')
        self.sub_font = Font(color='FF3B30', bold=True, size=10, name='Microsoft YaHei')
        self.thin_border = Border(
            left=Side(style='thin', color='E5E5EA'),
            right=Side(style='thin', color='E5E5EA'),
            top=Side(style='thin', color='E5E5EA'),
            bottom=Side(style='thin', color='E5E5EA')
        )
        self.alt_fill = PatternFill(start_color='F2F2F7', end_color='F2F2F7', fill_type='solid')

    def _style_header(self, cell):
        cell.fill = self.header_fill
        cell.font = self.header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = self.thin_border

    def _style_cell(self, cell, category=None, acos_value=None, alt=False):
        cell.border = self.thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.font = self.normal_font
        if alt:
            cell.fill = self.alt_fill
        if category == '加法':
            cell.font = self.add_font
        elif category == '减法':
            cell.font = self.sub_font
        if acos_value is not None:
            if acos_value > 60:
                cell.font = self.high_acos_font
            elif acos_value < 20:
                cell.font = self.low_acos_font

    def export_report(self, report_data: dict, output_path: str) -> str:
        wb = Workbook()

        # Sheet 1: 概览
        ws_overview = wb.active
        ws_overview.title = '概览'
        self._build_overview(ws_overview, report_data)

        # Sheet 2: 核心诊断
        ws_diag = wb.create_sheet('核心诊断')
        self._build_diagnosis(ws_diag, report_data)

        # Sheet 3: 流量结构树
        ws_tree = wb.create_sheet('流量结构树')
        self._build_traffic_tree(ws_tree, report_data.get('traffic_tree', []))

        # Sheet 4: 12层动作表（按广告活动分组）
        ws_actions = wb.create_sheet('12层动作表')
        self._build_actions(ws_actions, report_data.get('actions', []))

        # Sheet 5: 今日执行清单-加法
        ws_add = wb.create_sheet('今日执行清单-加法')
        self._build_today_list_add(ws_add, report_data.get('actions', []))

        # Sheet 6: 今日执行清单-减法
        ws_sub = wb.create_sheet('今日执行清单-减法')
        self._build_today_list_sub(ws_sub, report_data.get('actions', []))

        # Sheet 7: 7天监控计划
        ws_monitor = wb.create_sheet('7天监控计划')
        self._build_monitor_plan(ws_monitor, report_data.get('monitor_plan', []))

        # Sheet 8: 三套方案
        ws_plans = wb.create_sheet('三套方案')
        self._build_three_plans(ws_plans, report_data.get('three_plans', {}))

        wb.save(output_path)
        return output_path

    def _build_overview(self, ws, data):
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = 'Leo Ads Master - 12维度分析报告'
        title_cell.font = self.title_font
        title_cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 40

        ws.merge_cells('A2:H2')
        sub = ws['A2']
        sub.value = f"ASIN: {data.get('asin', '')}  |  分析师: {data.get('analyst', '')}  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        sub.font = self.subtitle_font
        sub.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        sub.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 25

        info_rows = [
            ['目标ACOS', f"{data.get('target_acos', 25)}%", '目标TACOS', f"{data.get('target_tacos', 12)}%", '', '', '', ''],
            ['SP ACOS', f"{data.get('acos', 0):.1f}%", 'TACOS', f"{data.get('tacos', 0):.1f}%", '', '', '', ''],
            ['数据完整度', data.get('data_completeness', '高'), '建议可信度', data.get('confidence', '高'), '', '', '', ''],
        ]
        for i, row in enumerate(info_rows, start=4):
            for j, val in enumerate(row, start=1):
                cell = ws.cell(row=i, column=j, value=val)
                cell.border = self.thin_border
                cell.font = self.normal_font
                if j in (1, 3, 5):
                    cell.font = Font(bold=True, size=10, name='Microsoft YaHei')
                    cell.fill = PatternFill(start_color='F2F2F7', end_color='F2F2F7', fill_type='solid')

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20

    def _build_diagnosis(self, ws, data):
        ws.merge_cells('A1:D1')
        ws['A1'] = 'SECTION 1: 核心诊断'
        ws['A1'].font = Font(bold=True, size=14, color='FF6B6B', name='Microsoft YaHei')
        ws.row_dimensions[1].height = 30

        diagnosis = data.get('diagnosis', {})
        rows = [
            ['主问题分类', diagnosis.get('main_issue', ''), '', ''],
            ['次问题分类', ', '.join(diagnosis.get('sub_issues', [])), '', ''],
            ['页面是否优先', '是' if diagnosis.get('page_first') else '否', '', ''],
            ['页面优先理由', diagnosis.get('page_first_msg', ''), '', ''],
            ['数据完整度', data.get('data_completeness', ''), '', ''],
            ['建议可信度', data.get('confidence', '高'), '', ''],
        ]
        for i, row in enumerate(rows, start=3):
            for j, val in enumerate(row, start=1):
                cell = ws.cell(row=i, column=j, value=val)
                cell.border = self.thin_border
                cell.font = self.normal_font
                if j == 1:
                    cell.font = Font(bold=True, size=10, name='Microsoft YaHei')
                    cell.fill = PatternFill(start_color='F2F2F7', end_color='F2F2F7', fill_type='solid')

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 60

    def _build_traffic_tree(self, ws, tree):
        ws.merge_cells('A1:I1')
        ws['A1'] = '流量结构树（按搜索量分层排序）'
        ws['A1'].font = Font(bold=True, size=14, color='FF6B6B', name='Microsoft YaHei')
        ws.row_dimensions[1].height = 30

        headers = ['层级', '关键词', '来源', '意图层级', '月搜索量', 'ABA排名', 'ACOS', '操作动作', '订单/点击']
        ws.append(headers)
        for cell in ws[3]:
            self._style_header(cell)

        level_colors = {
            '核心词': 'FFD6D6', '一级大词': 'FFE5CC', '二级大词': 'FFF2CC',
            '一级长尾词': 'E2EFDA', '二级长尾词': 'D9E1F2',
            '词根词': 'E1D5E7', '品牌词': 'D5E8D4', '活动词': 'F8CECC'
        }

        for item in tree:
            row = [
                item.get('level', ''),
                item.get('keyword', ''),
                item.get('source', ''),
                item.get('intent', ''),
                item.get('search_volume', ''),
                item.get('aba_rank', ''),
                item.get('acos', ''),
                item.get('action', ''),
                f"{item.get('orders', 0)} / {item.get('clicks', 0)}"
            ]
            ws.append(row)
            row_idx = ws.max_row
            level = item.get('level', '')
            fill_color = level_colors.get(level, 'FFFFFF')
            for col_idx in range(1, 10):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = self.thin_border
                cell.font = self.normal_font
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 12

    def _build_actions(self, ws, actions):
        ws.merge_cells('A1:M1')
        ws['A1'] = '12层动作表（按广告活动分组，ACOS最高组优先）'
        ws['A1'].font = Font(bold=True, size=14, color='FF6B6B', name='Microsoft YaHei')
        ws.row_dimensions[1].height = 30

        headers = ['广告活动', '广告组', '优先级', '维度层', '加减法', '广告类型', '目标词/ASIN',
                   '当前数据', '建议动作', '调整值', '原因', '预期影响', '执行时间']
        ws.append(headers)
        for cell in ws[3]:
            self._style_header(cell)
        ws.row_dimensions[3].height = 25

        prev_campaign = None
        for action in actions:
            is_new_campaign = action.get('campaign') != prev_campaign and prev_campaign is not None
            if is_new_campaign:
                # Insert blank separator row between campaigns
                ws.append([''] * 13)
                for col_idx in range(1, 14):
                    ws.cell(row=ws.max_row, column=col_idx).fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            prev_campaign = action.get('campaign')

            row = [
                action.get('campaign', ''),
                action.get('ad_group', ''),
                action.get('priority', ''),
                action.get('layer', ''),
                action.get('category', ''),
                action.get('ad_type', ''),
                action.get('target', ''),
                action.get('current_data', ''),
                action.get('action', ''),
                action.get('adjust_value', ''),
                action.get('reason', ''),
                action.get('expected_impact', ''),
                action.get('exec_time', '')
            ]
            ws.append(row)
            row_idx = ws.max_row
            for col_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                self._style_cell(cell, category=action.get('category'))

        col_widths = [28, 20, 8, 20, 8, 10, 22, 28, 14, 24, 35, 20, 10]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _build_today_list_add(self, ws, actions):
        ws.merge_cells('A1:D1')
        ws['A1'] = '今日执行清单 - 加法（扩量/加预算/加竞价/加词/加企业购）'
        ws['A1'].font = Font(bold=True, size=14, color='007AFF', name='Microsoft YaHei')
        ws.row_dimensions[1].height = 30

        headers = ['序号', '广告活动 / 广告组', '执行动作', '调整值']
        ws.append(headers)
        for cell in ws[3]:
            self._style_header(cell)

        add_actions = [a for a in actions if a.get('category') == '加法']
        for i, action in enumerate(add_actions, start=1):
            ws.append([
                i,
                f"{action.get('campaign', '')} / {action.get('ad_group', '')}",
                f"[{action.get('layer', '')}] {action.get('action', '')}",
                action.get('adjust_value', '')
            ])
            row_idx = ws.max_row
            for col_idx in range(1, 5):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = self.thin_border
                cell.font = self.add_font if col_idx >= 3 else self.normal_font
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 35

    def _build_today_list_sub(self, ws, actions):
        ws.merge_cells('A1:D1')
        ws['A1'] = '今日执行清单 - 减法（控预算/降竞价/否词/减品牌词）'
        ws['A1'].font = Font(bold=True, size=14, color='FF3B30', name='Microsoft YaHei')
        ws.row_dimensions[1].height = 30

        headers = ['序号', '广告活动 / 广告组', '执行动作', '调整值']
        ws.append(headers)
        for cell in ws[3]:
            self._style_header(cell)

        sub_actions = [a for a in actions if a.get('category') == '减法']
        for i, action in enumerate(sub_actions, start=1):
            ws.append([
                i,
                f"{action.get('campaign', '')} / {action.get('ad_group', '')}",
                f"[{action.get('layer', '')}] {action.get('action', '')}",
                action.get('adjust_value', '')
            ])
            row_idx = ws.max_row
            for col_idx in range(1, 5):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = self.thin_border
                cell.font = self.sub_font if col_idx >= 3 else self.normal_font
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 35

    def _build_monitor_plan(self, ws, plan):
        ws.merge_cells('A1:E1')
        ws['A1'] = 'SECTION 4: 7天监控计划'
        ws['A1'].font = Font(bold=True, size=14, color='FF6B6B', name='Microsoft YaHei')
        ws.row_dimensions[1].height = 30

        headers = ['监控目标', '监控指标', '阈值/Trigger', '触发动作', '时间窗口']
        ws.append(headers)
        for cell in ws[3]:
            self._style_header(cell)

        for item in plan:
            ws.append([
                item.get('target', ''),
                item.get('metric', ''),
                item.get('threshold', ''),
                item.get('trigger', ''),
                item.get('window', '')
            ])
            row_idx = ws.max_row
            for col_idx in range(1, 6):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = self.thin_border
                cell.font = self.normal_font
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 12

    def _build_three_plans(self, ws, plans):
        ws.merge_cells('A1:G1')
        ws['A1'] = '三套方案对比（保守 / 平衡 / 激进）'
        ws['A1'].font = Font(bold=True, size=14, color='FF6B6B', name='Microsoft YaHei')
        ws.row_dimensions[1].height = 30

        headers = ['方案', '预算调整', '竞价策略', '预估单量增幅', 'ACOS波动区间', '回落周期', '适用情况']
        ws.append(headers)
        for cell in ws[3]:
            self._style_header(cell)

        plan_colors = {'conservative': 'E2EFDA', 'balanced': 'D9E1F2', 'aggressive': 'FFD6D6'}
        for key, name in [('conservative', '保守'), ('balanced', '平衡'), ('aggressive', '激进')]:
            p = plans.get(key, {})
            ws.append([
                name,
                p.get('budget_change', ''),
                p.get('bid_strategy', ''),
                p.get('estimated_order_increase', ''),
                p.get('acos_range', ''),
                p.get('fallback_period', ''),
                p.get('suitable', '')
            ])
            row_idx = ws.max_row
            fill_color = plan_colors.get(key, 'FFFFFF')
            for col_idx in range(1, 8):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = self.thin_border
                cell.font = self.normal_font
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 28
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 28
